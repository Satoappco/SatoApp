#!/usr/bin/env python3
"""
Helper script to create a sample Q&A testing Excel template.

This creates a properly formatted Excel file with sample questions
that can be used with the test_qa_evaluation.py script.
"""

import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_template(output_path: str, with_samples: bool = True):
    """
    Create a Q&A testing Excel template.

    Args:
        output_path: Path where to save the Excel file
        with_samples: If True, include sample questions
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Tests"

    # Set column widths
    ws.column_dimensions['A'].width = 50  # Type
    ws.column_dimensions['B'].width = 50  # Question
    ws.column_dimensions['C'].width = 50  # Expected Answer
    ws.column_dimensions['D'].width = 50  # Current Answer
    ws.column_dimensions['E'].width = 50  # Previous Answer
    ws.column_dimensions['F'].width = 18  # Rank
    ws.column_dimensions['G'].width = 50  # Suggestion

    # Create header row with styling
    headers = [
        "Type",
        "Question",
        "Expected Answer",
        "Current Answer",
        "Previous Answer",
        "Rank",
        "Suggestion"
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Freeze the header row
    ws.freeze_panes = 'A2'

    if with_samples:
        # Add sample questions
        samples = [
            {
                "question": "What was our total Facebook Ads spend last month?",
                "expected": "A response that includes the total spend amount for Facebook Ads from the previous month, with currency formatting and comparison to previous period if available."
            },
            {
                "question": "Show me the top performing campaigns from Google Ads this week",
                "expected": "A list of the best performing Google Ads campaigns from the current week, including metrics like CTR, conversions, and cost per conversion."
            },
            {
                "question": "Compare Facebook and Google Ads performance for Q3",
                "expected": "A comparative analysis of Facebook Ads vs Google Ads for Q3, showing key metrics like spend, impressions, clicks, conversions, and ROI for both platforms."
            },
            {
                "question": "What are the trends in our click-through rates over the past 6 months?",
                "expected": "A trend analysis of CTR across all platforms over the last 6 months, identifying patterns, seasonality, or notable changes."
            },
            {
                "question": "Which campaigns need optimization based on their ROAS?",
                "expected": "A list of campaigns with poor ROAS (Return on Ad Spend) that need attention, including specific recommendations for optimization."
            },
        ]

        for row, sample in enumerate(samples, 2):
            # Question
            cell = ws.cell(row, 2)
            cell.value = sample["question"]
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Expected Answer
            cell = ws.cell(row, 3)
            cell.value = sample["expected"]
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Leave Current Answer, Previous Answer, Rank, and Suggestion empty
            for col in [4, 5, 6, 7]:
                cell = ws.cell(row, col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Save the workbook
    output = Path(output_path)
    wb.save(output)

    print(f"✅ Created template: {output}")
    print(f"   Rows: {len(samples) if with_samples else 0} sample questions")
    print(f"\nYou can now:")
    print(f"1. Add more questions to the Excel file")
    print(f"2. Run the evaluation: python test_qa_evaluation.py {output}")


def main():
    parser = argparse.ArgumentParser(
        description="Create a Q&A testing Excel template"
    )

    parser.add_argument(
        "output_file",
        nargs='?',
        default="qa_tests.xlsx",
        help="Output Excel file path (default: qa_tests.xlsx)"
    )

    parser.add_argument(
        "--no-samples",
        action="store_true",
        help="Create empty template without sample questions"
    )

    args = parser.parse_args()

    create_template(
        output_path=args.output_file,
        with_samples=not args.no_samples
    )


if __name__ == "__main__":
    main()
